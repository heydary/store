import openpyxl
from openpyxl import Workbook



wb1 = openpyxl.load_workbook("f:/store/Inventory.xlsx")
sheet1 = wb1.get_sheet_by_name('گزارش موجودی انبار')
row_count = sheet1.max_row
columns = 6




wb2 = openpyxl.load_workbook("f:/store/Ref.xlsx")
sheet2 = wb2.get_sheet_by_name('گزارش موجودی انبار')
row_count = sheet2.max_row
columns = 5






wb3 = Workbook()
sheet3 = wb3.active
wkb3_LastRow = sheet3.max_row




sheet1 = wb1.active
wkb1_LastRow = sheet1.max_row
sheet2 = wb2.active
wkb2_LastRow = sheet2.max_row
zrow=1




for xrow in range (2,(wkb1_LastRow+1)):
    for yrow in range (2,(wkb2_LastRow+1)):
      # for zrow in range(1,(wkb3_LastRow+1)):
        if sheet1.cell(row=xrow, column=4).value == sheet2.cell(row=yrow, column=4).value:
            if(sheet1.cell(row=xrow, column=6).value <= sheet2.cell(row=yrow, column=5).value) & (abs(sheet1.cell(row=xrow,column=6).value-sheet2.cell(row=xrow,column=5).value)>0):
               #print(('%-18s'%(sheet2.cell(row=xrow,column=5).value-sheet1.cell(row=xrow,column=6).value)),(sheet1.cell(row=xrow, column=4).value))
               sheet3.cell(row=zrow, column=1).value=('%-18s'%(sheet1.cell(row=xrow,column=6).value - sheet2.cell(row=xrow,column=5).value))
               sheet3.cell(row=zrow, column=1).value=(sheet1.cell(row=xrow, column=6).value - sheet2.cell(row=xrow, column=5).value)
               sheet3.cell(row=zrow, column=2).value=sheet1.cell(row=xrow, column=4).value
               zrow+=1



wb3.save("f:/store/Report.xlsx")